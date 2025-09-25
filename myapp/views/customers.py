# views.py
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

from myapp.models import Customer
from myapp.serializers import CustomerSerializer
from myapp.permissions import IsAdmin

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        "area", "zone_number", "plot_number", "property_type", "account_number"
    ]  
    search_fields = ["full_name", "phone", "location_link"]
    ordering_fields = ["id", "full_name", "account_number", "starting_date"]
    ordering = ["id"]
    
    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Customers"

        headers = [
            "ID", "Full Name", "Phone", "Area", "Zone Number", "Plot Number",
            "Property Type", "Account Number", "Starting Date",
            "Agreement Without Meter", "Weekly Trips","Gallons", 
            "Filling Stations", "Location Link", "Delivery Days", 
            "Driver Username", "Delivery Time"
        ]
        sheet.append(headers)

        for customer in queryset:
            sheet.append([
                customer.id,
                customer.full_name,
                customer.phone,
                customer.area,
                customer.zone_number,
                customer.plot_number,
                customer.property_type,
                customer.account_number,
                customer.starting_date.strftime("%Y-%m-%d") if customer.starting_date else "",
                customer.agreement_without_meter,
                customer.weekly_trips,
                customer.gallons,
                customer.filling_stations,
                customer.location_link,
                ", ".join(customer.delivery_days or []),
                customer.driver.username if customer.driver else "",
                customer.delivery_time,
            ])


        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="customers.xlsx"'
        workbook.save(response)
        return response