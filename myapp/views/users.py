from rest_framework import viewsets, filters
from myapp.models import User
from myapp.serializers import UserSerializer
from myapp.permissions import IsAdmin
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['role']  
    search_fields = ['username', 'first_name', 'last_name', 'phone', 'email']
    ordering_fields = ['id', 'username', 'role']  
    ordering = ['id']  

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        headers = ["ID", "Username", "First Name", "Last Name", "Email", "Role", "Phone", "Date Joined", "Last Login"]
        sheet.append(headers)

        for user in queryset:
            sheet.append([
				user.id,
				user.username,
				user.first_name,
				user.last_name,
				user.email,
				user.role,
				user.phone,
				user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else "",
				user.last_login.strftime("%Y-%m-%d %H:%M") if user.last_login else "",
			])

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(
			content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
		)
        response["Content-Disposition"] = 'attachment; filename="users.xlsx"'
        workbook.save(response)
        return response
