from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from myapp.permissions import IsAdmin, IsAccountant
from django.http import HttpResponse
import openpyxl
from myapp.models import Order, User, Customer, Invoice

class ExportOrdersExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAccountant]

    def get(self, request, format=None):
    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"

        
        headers = ["ID", "Customer", "Driver", "Status", "Created At", "Confirmed At"]
        ws.append(headers)

        
        orders = Order.objects.all() 
        for order in orders:
            ws.append([
                order.id,
                order.customer.full_name if order.customer else "",
                order.driver.username if order.driver else "",
                order.status,
                order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
                order.confirmed_at.strftime("%Y-%m-%d %H:%M") if order.confirmed_at else ""
            ])


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        wb.save(response)
        return response

class ExportUsersExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        # العناوين
        headers = ["ID", "Username", "Email", "Role", "Phone", "Date Joined", "Last Login"]
        sheet.append(headers)

        # البيانات
        for user in User.objects.all():
            sheet.append([
                user.id,
                user.username,
                user.email,
                user.role,
                user.phone,
                user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else "",
                user.last_login.strftime("%Y-%m-%d %H:%M") if user.last_login else "",
            ])

        # تجهيز الملف
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="users.xlsx"'
        workbook.save(response)
        return response


class ExportCustomersExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Customers"

        # العناوين
        headers = [
            "ID", "Full Name", "Phone", "Area", "Zone Number", "Plot Number",
            "Property Type", "Account Number", "Starting Date",
            "Agreement Without Meter", "Weekly Trips", "Two Trips",
            "Three Trips", "Four Trips", "Gallons", "Filling Stations", "Location Link"
        ]
        sheet.append(headers)

        # البيانات
        for customer in Customer.objects.all():
            sheet.append([
                customer.id,
                customer.full_name,
                customer.phone,
                customer.area,
                customer.zone_number,
                customer.plot_number,
                customer.property_type,
                customer.account_number,
                customer.starting_date.strftime("%Y-%m-%d") if customer.starting_date else "",
                customer.agreement_without_meter,
                customer.weekly_trips,
                customer.two_trips,
                customer.three_trips,
                customer.four_trips,
                customer.gallons,
                customer.filling_stations,
                customer.location_link,
            ])

        # تجهيز الملف
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="customers.xlsx"'
        workbook.save(response)
        return response
    
class ExportInvoicesExcelView(APIView):
    permission_classes = [IsAuthenticated, (IsAdmin | IsAccountant)]

    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"

        headers = ["ID", "Customer", "Order ID", "Issued By", "Base Amount", "Tax %", "Discount", "Extra Fees", "Total", "Paid", "Due Date", "Created At"]
        sheet.append(headers)

        for inv in Invoice.objects.filter(order__status="confirmed"):
            sheet.append([
                inv.id,
                inv.order.customer.full_name,
                inv.order.id,
                inv.issued_by.username if inv.issued_by else "",
                inv.base_amount,
                inv.tax_percentage,
                inv.discount_amount,
                inv.extra_fees,
                inv.total_amount,
                inv.paid,
                inv.due_date.strftime("%Y-%m-%d"),
                inv.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="invoices.xlsx"'
        workbook.save(response)
        return response
    
#pdf
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from myapp.models import Invoice
from myapp.permissions import IsAdmin, IsAccountant

class ExportInvoicesPDFView(APIView):
    permission_classes = [IsAuthenticated, (IsAdmin | IsAccountant)]

    def get(self, request, *args, **kwargs):

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="invoices.pdf"'


        pdf = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        y = height - 50
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(50, y, "Invoices Report")
        y -= 30

        pdf.setFont("Helvetica-Bold", 12)
        headers = ["ID", "Customer", "Order ID", "Issued By", "Total", "Paid", "Due Date"]
        x_positions = [50, 100, 250, 320, 400, 460, 500]
        for i, header in enumerate(headers):
            pdf.drawString(x_positions[i], y, header)
        y -= 20

        pdf.setFont("Helvetica", 10)

        for inv in Invoice.objects.filter(order__status="confirmed"):
            if y < 50:
                pdf.showPage()
                y = height - 50
                pdf.setFont("Helvetica", 10)
            row = [
                str(inv.id),
                inv.order.customer.full_name,
                str(inv.order.id),
                inv.issued_by.username if inv.issued_by else "",
                f"{inv.total_amount:.2f}",
                "Yes" if inv.paid else "No",
                inv.due_date.strftime("%Y-%m-%d")
            ]
            for i, cell in enumerate(row):
                pdf.drawString(x_positions[i], y, cell)
            y -= 20

        pdf.showPage()
        pdf.save()
        return response
