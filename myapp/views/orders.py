# views.py
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

from myapp.models import Order
from myapp.serializers import OrderSerializer, DriverOrderSerializer
from myapp.permissions import IsAdmin, IsDriver

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status", "customer__full_name", "driver__username"]
    search_fields = ["customer__full_name", "driver__username"]
    ordering_fields = ["created_at", "confirmed_at", "status"]
    ordering = ["-created_at"]
    
    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Orders"

        headers = [
            "ID", "Customer", "Driver", "Status",
            "Created At", "Confirmed At", "Filled Amount",
            "Is Late", "Problem Reason", "Proof Image"
        ]
        sheet.append(headers)

        for order in queryset:
            sheet.append([
                order.id,
                order.customer.full_name if order.customer else "-",
                order.driver.username if order.driver else "-",
                order.status,
                order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
                order.confirmed_at.strftime("%Y-%m-%d %H:%M") if order.confirmed_at else "",
                getattr(order, "filled_amount", "-"),
                order.is_driver_late(minutes=30),
                getattr(order, "problem_reason", "-"),
                order.proof_image.url if order.proof_image else "-",
            ])

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
        workbook.save(response)
        return response


# from openpyxl.drawing.image import Image as ExcelImage
# import os
# from django.http import HttpResponse
# 
# @action(detail=False, methods=['get'])
# def export_excel_with_images(self, request):
#     queryset = self.filter_queryset(self.get_queryset())
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Orders"
#
#     headers = [
#         "ID", "Customer", "Driver", "Status",
#         "Created At", "Confirmed At", "Filled Amount",
#         "Is Late", "Problem Reason", "Proof Image"
#     ]
#     sheet.append(headers)
#
#     for order in queryset:
#         row = [
#             order.id,
#             order.customer.full_name if order.customer else "-",
#             order.driver.username if order.driver else "-",
#             order.status,
#             order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
#             order.confirmed_at.strftime("%Y-%m-%d %H:%M") if order.confirmed_at else "",
#             getattr(order, "filled_amount", "-"),
#             order.is_driver_late(minutes=30),
#             getattr(order, "problem_reason", "-"),
#             ""  # proof image placeholder
#         ]
#         sheet.append(row)
#
#         if order.proof_image and os.path.exists(order.proof_image.path):
#             img = ExcelImage(order.proof_image.path)
#             img.width, img.height = 80, 80
#             sheet.add_image(img, f"J{sheet.max_row}")
#
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="orders_with_images.xlsx"'
#     workbook.save(response)
#     return response


class DriverOrderViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = DriverOrderSerializer
    permission_classes = [IsAuthenticated, IsDriver]
    queryset = Order.objects.none()

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status"]
    search_fields = ["customer__full_name"]
    ordering_fields = ["created_at", "confirmed_at"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return Order.objects.filter(driver=self.request.user)

    @action(detail=True, methods=["post"])
    def confirm(self, request, pk=None):
        order = self.get_object()
        if order.driver != request.user:
            return Response({"error": "This order is not assigned to you"}, status=403)
        if order.status != "pending":
            return Response({"error": "Order already confirmed or completed"}, status=400)

        filled_amount = request.data.get("filled_amount")
        proof_image = request.FILES.get("proof_image")

        if not filled_amount or not proof_image:
            return Response({"error": "Both filled amount and proof image are required"}, status=400)

        order.confirm(filled_amount=filled_amount, proof_image=proof_image)
        serializer = self.get_serializer(order)
        return Response({
            "message": "Order confirmed successfully",
            "confirmed_at": order.confirmed_at,
            "order": serializer.data
        }, status=200)

    @action(detail=True, methods=["post"])
    def problem(self, request, pk=None):
        order = self.get_object()
        if order.driver != request.user:
            return Response({"error": "This order is not assigned to you"}, status=403)

        reason = request.data.get("reason")
        if not reason:
            return Response({"error": "Problem reason is required"}, status=400)

        order.mark_problem(reason)
        serializer = self.get_serializer(order)
        return Response({
            "message": "Problem reported",
            "order": serializer.data
        }, status=200)
